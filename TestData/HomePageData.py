import openpyxl


class HomePageData:

    # test_HomePage_data = [{"firstname":"Mike", "email":"test@practice.com", "pass":"abc123!", "gender":"Male"}, {"firstname":"Maria", "email":"test@practice.com", "pass":"abc", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        data_dict = {}
        book = openpyxl.load_workbook("C:\\Users\\destr\\OneDrive\\Desktop\\Python Training\\PythonPractice.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:  # to only get column 2
                for j in range(2, sheet.max_column + 1):  # to get columns
                # dict["lastname"] = "McDonald"
                    data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[data_dict]